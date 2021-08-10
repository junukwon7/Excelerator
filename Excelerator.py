import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import numpy as np
from PIL import Image;

print('Input Image Filename: ')
image_file = input()
print('Output Excel Filename: ')
excel_file = input()

def rgb2hex(r, g, b):
	return '{:02x}{:02x}{:02x}'.format(r, g, b, )

im = Image.open(image_file)
h, w = im.size
im = im.convert('RGB')
mat = np.array(im)

wb = Workbook()
sheet = wb.active


for col_cnt in range(0, h):
	print ('Finished {} / {} columns. {}% Complete'.format(col_cnt, h, round(col_cnt / h * 100)),  end="\r")
	sheet.column_dimensions[get_column_letter(col_cnt+1)].width = 2.7
	for row_cnt in range(0, w):
		code = rgb2hex(mat[row_cnt][col_cnt][0], mat[row_cnt][col_cnt][1], mat[row_cnt][col_cnt][2])
		sheet.cell(row = row_cnt+1, column = col_cnt+1).fill = PatternFill(start_color = code, end_color = code, fill_type = "solid")

wb.save(filename = excel_file)

print('Exceleration Complete! Saved as: ' + excel_file)